#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Process trade printing Excel orders and append grouped rows to Google Sheet via gog.

Handles variant header where column 'thời gian' is inserted after 'Tên hạng mục'.

Outputs a JSON summary to stdout.
"""

import argparse
import json
import math
import os
import re
import subprocess
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except Exception as e:
    print("ERROR: openpyxl import failed. Install with: pip install openpyxl", file=sys.stderr)
    raise

CODES_ORDER = [
    "BVI","CNG","HDI","CSO","LQU","PXU","XKH","DAN","TTI","XMA","PYE","TNG","CLI","KMO","TNU",
    "TTH","VYE","TDI","PHY","GTH","DTR","QYE","DVA","LNH","VP",
]
CODE_TO_INDEX = {c: i for i, c in enumerate(CODES_ORDER)}  # 0..24 map to J..AH

SHEET_ID = "1cHzgteJN6LfhrQdiZI8IABp_rJ17Xr-RoAqOoTCCBWc"
TAB_NAME = "Tổng hợp"


def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s


def norm_header(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    s = strip_accents(s).lower()
    s = re.sub(r"\s+", " ", s)
    return s


def parse_label_value(cell_value, label_prefix):
    """Parse cases like 'NCC: ABC' or separate cells.

    Returns value after prefix if present, else None.
    """
    if cell_value is None:
        return None
    s = str(cell_value).strip()
    if not s:
        return None
    # normalize colon variants
    if s.lower().startswith(label_prefix.lower()):
        rest = s[len(label_prefix):].strip()
        if rest.startswith(":"):
            rest = rest[1:].strip()
        return rest or None
    return None


def pick_row_value(ws, row_idx, label_prefix):
    """Row contains something like ['NCC:', None, 'ABC'].

    Priority:
    - If any cell is like 'NCC: ABC' -> take ABC
    - Else take first non-empty string cell that is not just label ('NCC:'/'CT:')
    """
    cells = [ws.cell(row=row_idx, column=c).value for c in range(1, 10)]
    # Inline label
    for v in cells:
        parsed = parse_label_value(v, label_prefix)
        if parsed:
            return parsed

    # Otherwise, find a cell that's not the label
    label_norm = strip_accents(label_prefix).lower().rstrip(":")
    for v in cells:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        s_norm = strip_accents(s).lower().rstrip(":")
        if s_norm in (label_norm, label_norm + ":"):
            continue
        # if cell contains just 'NCC:'
        if s_norm.startswith(label_norm) and s_norm.replace(label_norm, "").strip() in (":", ""):
            continue
        return s

    return ""


def find_header_row(ws, target_header="Tên hạng mục", max_scan=20):
    target_norm = norm_header(target_header)
    for r in range(1, max_scan + 1):
        row_vals = [norm_header(ws.cell(row=r, column=c).value) for c in range(1, 25)]
        if target_norm in row_vals:
            return r
    raise ValueError("Could not find header row containing 'Tên hạng mục'")


def coerce_number(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
            return None
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    # remove thousand separators
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def parse_date_cell(x):
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    if not s:
        return None
    # common VN formats
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    # try parse with day first if available
    try:
        import dateutil.parser  # type: ignore
        dt = dateutil.parser.parse(s, dayfirst=True)
        return dt.date()
    except Exception:
        return None


def format_vn_date(d: date) -> str:
    return d.strftime("%d/%m/%Y")


def build_header_map(ws, header_row):
    header_map = {}
    for c in range(1, 60):
        hv = ws.cell(row=header_row, column=c).value
        if hv is None:
            continue
        key = norm_header(hv)
        if key:
            header_map[key] = c
    return header_map


def get_col(header_map, *names):
    for name in names:
        n = norm_header(name)
        if n in header_map:
            return header_map[n]
    return None


def process_one_file(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    ncc = pick_row_value(ws, 1, "NCC")
    ct = pick_row_value(ws, 2, "CT")

    header_row = find_header_row(ws, "Tên hạng mục", max_scan=30)
    header_map = build_header_map(ws, header_row)

    col_ten = get_col(header_map, "Tên hạng mục")
    col_time = get_col(header_map, "thời gian", "thoi gian")
    col_code = get_col(header_map, "Code", "Mã", "Ma")
    col_qty = get_col(header_map, "Số lượng", "So luong", "SL")
    col_chat = get_col(header_map, "Chất liệu", "Chat lieu")
    col_price = get_col(header_map, "Đơn giá", "Don gia")
    col_dai = get_col(header_map, "Dài", "Dai")
    col_rong = get_col(header_map, "Rộng", "Rong")
    col_amt = get_col(header_map, "Thành Tiền", "Thanh Tien", "Thành tiền")

    required = {
        "Tên hạng mục": col_ten,
        "Code": col_code,
        "Số lượng": col_qty,
        "Chất liệu": col_chat,
        "Đơn giá": col_price,
        "Dài": col_dai,
        "Rộng": col_rong,
        "Thành Tiền": col_amt,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        raise ValueError(f"Missing required columns in header row {header_row}: {missing}. Found headers={list(header_map.keys())}")

    start_row = header_row + 1

    groups = {}
    file_total_amt = 0.0
    file_total_qty = 0.0
    unknown_codes = Counter()
    all_dates = []

    def key_tuple(ten, chat, price, dai, rong):
        return (
            (ten or "").strip(),
            (chat or "").strip(),
            float(price or 0.0),
            float(dai or 0.0),
            float(rong or 0.0),
        )

    r = start_row
    empty_streak = 0
    while True:
        ten = ws.cell(row=r, column=col_ten).value
        code = ws.cell(row=r, column=col_code).value
        qty = ws.cell(row=r, column=col_qty).value
        chat = ws.cell(row=r, column=col_chat).value
        price = ws.cell(row=r, column=col_price).value
        dai = ws.cell(row=r, column=col_dai).value
        rong = ws.cell(row=r, column=col_rong).value
        amt = ws.cell(row=r, column=col_amt).value
        tval = ws.cell(row=r, column=col_time).value if col_time else None

        # Detect blank row: if all major fields empty, count empties and break at first blank block.
        if all(v in (None, "") for v in (ten, code, qty, chat, price, dai, rong, amt, tval)):
            empty_streak += 1
            if empty_streak >= 1:
                break
            r += 1
            continue
        empty_streak = 0

        ten_s = str(ten).strip() if ten is not None else ""
        code_s = str(code).strip().upper() if code is not None else ""
        chat_s = str(chat).strip() if chat is not None else ""
        qty_n = coerce_number(qty) or 0.0
        price_n = coerce_number(price) or 0.0
        dai_n = coerce_number(dai) or 0.0
        rong_n = coerce_number(rong) or 0.0
        amt_n = coerce_number(amt) or 0.0

        if col_time:
            d = parse_date_cell(tval)
            if d:
                all_dates.append(d)

        file_total_amt += amt_n
        file_total_qty += qty_n

        k = key_tuple(ten_s, chat_s, price_n, dai_n, rong_n)
        if k not in groups:
            groups[k] = {
                "ten": ten_s,
                "chat": chat_s,
                "price": price_n,
                "dai": dai_n,
                "rong": rong_n,
                "qty_total": 0.0,
                "amt_total": 0.0,
                "per_code": defaultdict(float),
            }
        g = groups[k]
        g["qty_total"] += qty_n
        g["amt_total"] += amt_n

        if code_s:
            if code_s in CODE_TO_INDEX:
                g["per_code"][code_s] += qty_n
            else:
                unknown_codes[code_s] += qty_n

        r += 1
        if r > ws.max_row + 5:
            break

    # Decide file date
    if all_dates:
        d = Counter(all_dates).most_common(1)[0][0]
    else:
        d = date(2026, 3, 3)  # per instruction: prefer date in file; fallback to 3/3/2026

    # Build gsheet rows A..AI
    out_rows = []
    qc1_all = True

    for k, g in groups.items():
        row = [""] * 35
        row[0] = format_vn_date(d)
        row[1] = ct
        row[2] = g["ten"]
        row[3] = g["chat"]
        row[4] = g["rong"]
        row[5] = g["dai"]
        row[6] = g["price"]
        row[7] = g["amt_total"]
        row[8] = g["qty_total"]
        # J..AH
        sum_codes = 0.0
        for code, q in g["per_code"].items():
            idx = CODE_TO_INDEX[code]
            row[9 + idx] = q
            sum_codes += q
        row[34] = ncc

        # QC1: qty_total equals sum J..AH AND no unknown codes contributing within group
        # (Unknown codes are tracked globally; if present, QC1 may fail for the file.)
        if abs(g["qty_total"] - sum_codes) > 1e-6:
            qc1_all = False
        out_rows.append(row)

    # QC2
    sum_group_amt = sum(g["amt_total"] for g in groups.values())
    qc2 = abs(sum_group_amt - file_total_amt) <= 1e-6

    result = {
        "file": xlsx_path,
        "ncc": ncc,
        "ct": ct,
        "header_row": header_row,
        "date": format_vn_date(d),
        "rows_to_append": len(out_rows),
        "groups": [
            {
                "ten": g["ten"],
                "chat": g["chat"],
                "price": g["price"],
                "dai": g["dai"],
                "rong": g["rong"],
                "qty": g["qty_total"],
                "amt": g["amt_total"],
                "per_code": dict(g["per_code"]),
            }
            for g in groups.values()
        ],
        "unknown_codes": dict(unknown_codes),
        "qc": {
            "qc1_qty_equals_sum_codes": qc1_all and (len(unknown_codes) == 0),
            "qc2_total_amt_matches_file": qc2,
            "file_total_amt": file_total_amt,
            "sum_group_amt": sum_group_amt,
            "file_total_qty": file_total_qty,
        },
        "out_rows": out_rows,
    }
    return result


def gog_append_rows(sheet_id, tab_name, rows):
    if not rows:
        return {"ok": True, "stdout": "(no rows)", "stderr": ""}

    cmd = [
        "gog",
        "sheets",
        "append",
        sheet_id,
        f"{tab_name}!A:AI",
        "--values-json",
        json.dumps(rows, ensure_ascii=False),
        "--insert",
        "INSERT_ROWS",
    ]

    env = os.environ.copy()
    p = subprocess.run(cmd, capture_output=True, text=True, env=env)
    return {
        "ok": p.returncode == 0,
        "returncode": p.returncode,
        "stdout": p.stdout.strip(),
        "stderr": p.stderr.strip(),
        "cmd": cmd,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("files", nargs="+")
    args = ap.parse_args()

    all_results = []
    for f in args.files:
        res = process_one_file(f)
        if not args.dry_run:
            append_res = gog_append_rows(SHEET_ID, TAB_NAME, res["out_rows"])
            res["append"] = append_res
        else:
            res["append"] = {"ok": True, "stdout": "DRY_RUN", "stderr": ""}
        # Remove out_rows from printing huge? keep but also
        all_results.append(res)

    print(json.dumps(all_results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
