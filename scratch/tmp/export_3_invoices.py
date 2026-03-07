import json, re, zipfile, shutil, os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

TEMPLATE = "/data/workspace/templates/Hanh_expense_template.xlsx"
OUT_DIR = "/data/workspace/out"
os.makedirs(OUT_DIR, exist_ok=True)

# Invoices in this batch
invoices = [
  {"date":"2025-05-17","description":"Cafe","expense_account":"vietnam","value":75000,"currency":"VND"},
  {"date":"2025-04-07","description":"Dinner with agent","expense_account":"cambodia","value":2.75,"currency":"USD"},
  {"date":"2025-04-06","description":"Lunch","expense_account":"cambodia","value":8.45,"currency":"USD"},
]

# Copy template to work file
stamp = datetime.now(ZoneInfo("Asia/Ho_Chi_Minh")).strftime("%Y%m%d_%H%M%S")
work_path = os.path.join(OUT_DIR, f"Expenses_export_demo_{stamp}.xlsx")
shutil.copy(TEMPLATE, work_path)

wb = load_workbook(work_path)
ws = wb["Month_Year"]

# Mapping table C45:D49
code_map = {}
for r in range(45, 50):
    k = ws[f"C{r}"].value
    v = ws[f"D{r}"].value
    if k and v:
        code_map[str(k).strip().lower()] = str(v).strip()

sea_code = code_map.get("sea", "8741-4101")

# Rates in A55-A57 like: "1 THB = 840 VND"
rate_re = re.compile(r"=\s*(\d+(?:[\.,]\d+)?)\s*VND", re.IGNORECASE)

def parse_rate(cell):
    if not cell:
        return None
    m = rate_re.search(str(cell))
    if not m:
        return None
    s = m.group(1).replace(",", ".")
    try:
        return float(s)
    except:
        return None

rate_thb = parse_rate(ws["A55"].value)
rate_usd = parse_rate(ws["A56"].value)
rate_khr = parse_rate(ws["A57"].value)

rates = {"THB": rate_thb, "USD": rate_usd, "KHR": rate_khr}

# Fill rows starting from 5
start_row = 5
for idx, inv in enumerate(invoices, start=1):
    r = start_row + (idx - 1)

    # Account code
    ea = inv["expense_account"].lower().strip()
    if ea in ("laos", "other"):
        acct_code = sea_code
    else:
        acct_code = code_map.get(ea, sea_code)

    # Amount in VND
    cur = inv["currency"].upper().strip()
    val = inv["value"]
    if cur == "VND":
        amount_vnd = float(val)
    else:
        rate = rates.get(cur)
        if rate is None:
            raise RuntimeError(f"Missing FX rate for {cur} in template (A55-A57)")
        amount_vnd = float(val) * float(rate)

    # Write cells (ONLY A..F)
    ws[f"A{r}"].value = idx
    ws[f"B{r}"].value = datetime.fromisoformat(inv["date"])
    ws[f"B{r}"].number_format = "dd/mm/yyyy"
    ws[f"C{r}"].value = inv["description"]
    ws[f"D{r}"].value = acct_code
    ws[f"E{r}"].value = amount_vnd
    ws[f"F{r}"].value = None  # keep blank

wb.save(work_path)

# Patch externalLink artifacts from template to avoid Excel repair prompt
restore_paths = [
    "xl/externalLinks/externalLink1.xml",
    "xl/externalLinks/_rels/externalLink1.xml.rels",
]

with zipfile.ZipFile(TEMPLATE, "r") as zsrc:
    restore_bytes = {p: zsrc.read(p) for p in restore_paths}

patched_path = os.path.join(OUT_DIR, f"Expenses_export_demo_{stamp}_PERFECT.xlsx")
with zipfile.ZipFile(work_path, "r") as zin, zipfile.ZipFile(patched_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
    existing = set(zin.namelist())
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename in restore_bytes:
            data = restore_bytes[item.filename]
        zout.writestr(item, data)
    for p, b in restore_bytes.items():
        if p not in existing:
            zout.writestr(p, b)

os.remove(work_path)
print(patched_path)
