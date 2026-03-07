from openpyxl import load_workbook
import json

file_path = '/data/.openclaw/media/inbound/file_16---6ea47fea-67c8-40af-9ece-40d121d6a274.xlsx'

wb = load_workbook(file_path, data_only=True)
ws = wb.active

# Parse A1 and A2
A1_raw = str(ws['A1'].value)
A2_raw = str(ws['A2'].value)
A1_value = A1_raw.split(': ', 1)[1] if ': ' in A1_raw else A1_raw
A2_value = A2_raw.split(': ', 1)[1] if ': ' in A2_raw else A2_raw

# Column indices based on Excel (A=1, B=2, ..., H=8, I=9)
# Header row 3: C=Tên hạng mục, D=Chất liệu, E=Dài, F=Rộng, G=Đơn giá, H=Thành tiền, I=SL
data = {}
for row in range(5, ws.max_row + 1):
    name = ws.cell(row=row, column=3).value
    mat = ws.cell(row=row, column=4).value
    l = ws.cell(row=row, column=5).value
    w = ws.cell(row=row, column=6).value
    price = ws.cell(row=row, column=7).value
    total = ws.cell(row=row, column=8).value
    qty = ws.cell(row=row, column=9).value
    
    if name is None: continue
    
    key = (name, mat, price, l, w)
    if key not in data:
        data[key] = {'qty': 0, 'total': 0}
    data[key]['qty'] += (qty or 0)
    data[key]['total'] += (total or 0)

# Notes mapping: ncc mapping for 3 items
# Items 1 & 2: Đối tác vàng
# Item 3: Hoàng Việt
items = list(data.keys())
final_rows = []
for i, (key, vals) in enumerate(data.items()):
    ncc = "Đối tác vàng" if i < 2 else "Hoàng Việt"
    row = [
        '26/02/2026', # A
        A2_value,     # B
        key[0],       # C (Name)
        key[1],       # D (Material)
        key[4],       # E (Width - Rộng)
        key[3],       # F (Length - Dài)
        key[2],       # G (Price)
        vals['total'],# H
        vals['qty'],  # I
        *([None]*25), # J-AH (25 columns)
        ncc           # AI
    ]
    final_rows.append(row)

print(json.dumps(final_rows))